# apps/comptabilite/services/export_service.py
import csv
import io
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os


class ExportComptableService:
    """Service d'export comptable (PDF, Excel, CSV)"""
    
    @staticmethod
    def export_balance_csv(balance_data, total_debit, total_credit):
        """Export balance en CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Code compte', 'Libellé', 'Total Débit (F)', 'Total Crédit (F)', 'Solde (F)'])
        
        for ligne in balance_data:
            writer.writerow([
                ligne['compte__code'],
                ligne['compte__libelle'],
                f"{ligne['total_debit']:,.0f}",
                f"{ligne['total_credit']:,.0f}",
                f"{ligne['solde']:,.0f}"
            ])
        
        writer.writerow([])
        writer.writerow(['TOTAUX GÉNÉRAUX', '', f"{total_debit:,.0f}", f"{total_credit:,.0f}", f"{total_debit - total_credit:,.0f}"])
        
        return response
    
    @staticmethod
    def export_balance_excel(balance_data, total_debit, total_credit, exercice):
        """Export balance en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Balance générale"
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="0F3B2C", end_color="0F3B2C", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête
        ws.merge_cells('A1:E1')
        ws['A1'] = f"BALANCE GÉNÉRALE - Exercice {exercice.code}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Arrêté au {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes colonnes
        headers = ['Code compte', 'Libellé', 'Total Débit (F)', 'Total Crédit (F)', 'Solde (F)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Données
        for row, ligne in enumerate(balance_data, 5):
            ws.cell(row=row, column=1, value=ligne['compte__code']).border = thin_border
            ws.cell(row=row, column=2, value=ligne['compte__libelle']).border = thin_border
            ws.cell(row=row, column=3, value=float(ligne['total_debit'])).border = thin_border
            ws.cell(row=row, column=4, value=float(ligne['total_credit'])).border = thin_border
            ws.cell(row=row, column=5, value=float(ligne['solde'])).border = thin_border
        
        # Totaux
        total_row = len(balance_data) + 5
        ws.cell(row=total_row, column=2, value="TOTAUX GÉNÉRAUX").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=float(total_debit)).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=float(total_credit)).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(total_debit - total_credit)).font = Font(bold=True)
        
        # Ajuster les largeurs
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def export_balance_pdf(balance_data, total_debit, total_credit, exercice):
        """Export balance en PDF"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, alignment=1, spaceAfter=20)
        elements.append(Paragraph(f"BALANCE GÉNÉRALE - Exercice {exercice.code}", title_style))
        elements.append(Paragraph(f"Arrêté au {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Tableau
        data = [['Code', 'Libellé', 'Débit (F)', 'Crédit (F)', 'Solde (F)']]
        for ligne in balance_data[:50]:  # Limite à 50 lignes pour PDF
            data.append([
                ligne['compte__code'],
                ligne['compte__libelle'][:40],
                f"{ligne['total_debit']:,.0f}",
                f"{ligne['total_credit']:,.0f}",
                f"{ligne['solde']:,.0f}"
            ])
        
        if len(balance_data) > 50:
            data.append(['...', f"et {len(balance_data) - 50} autres comptes...", '', '', ''])
        
        data.append(['', 'TOTAUX GÉNÉRAUX', f"{total_debit:,.0f}", f"{total_credit:,.0f}", f"{total_debit - total_credit:,.0f}"])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F3B2C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
    
    @staticmethod
    def export_ecritures_csv(ecritures, total_debit, total_credit):
        """Export écritures en CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="ecritures_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Référence', 'Date', 'Journal', 'Libellé', 'Débit (F)', 'Crédit (F)', 'Validée'])
        
        for ecriture in ecritures:
            writer.writerow([
                ecriture.reference,
                ecriture.date_ecriture.strftime('%d/%m/%Y'),
                ecriture.journal.code,
                ecriture.libelle[:100],
                f"{ecriture.total_debit:,.0f}",
                f"{ecriture.total_credit:,.0f}",
                'Oui' if ecriture.validee else 'Non'
            ])
        
        writer.writerow([])
        writer.writerow(['TOTAUX', '', '', '', f"{total_debit:,.0f}", f"{total_credit:,.0f}", ''])
        
        return response
    
    
    